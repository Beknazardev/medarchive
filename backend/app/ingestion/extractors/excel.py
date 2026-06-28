"""Excel extractor - reads XLSX files using openpyxl in read-only, data-only mode."""

from __future__ import annotations

import time
from io import BytesIO
from typing import Any

import openpyxl
from openpyxl.utils import get_column_letter

from app.ingestion.extractors.base import (
    ExtractionLimits,
    ExtractionOutput,
    GenericCell,
    GenericRow,
    GenericTable,
    GenericTextBlock,
    MalformedDocument,
    UnsupportedFormat,
    check_timeout,
)

MAX_XLSX_ROWS = 1_048_576
MAX_XLSX_COLS = 16_384


def extract_excel(
    content: bytes,
    *,
    source_url: str | None = None,
    limits: ExtractionLimits | None = None,
) -> ExtractionOutput:
    lim = limits or ExtractionLimits()
    start = time.monotonic()
    errors: list[Any] = []
    tables: list[GenericTable] = []
    text_blocks: list[GenericTextBlock] = []

    if len(content) < 100:
        raise MalformedDocument("Excel content too small", stage="excel")

    try:
        wb = openpyxl.load_workbook(
            BytesIO(content),
            read_only=True,
            data_only=True,
        )
    except Exception as exc:
        msg = str(exc).lower()
        if "password" in msg or "encrypted" in msg:
            raise MalformedDocument(f"Excel file requires password: {exc}", stage="excel") from exc
        if "xls" in msg and "xlsx" not in msg:
            raise UnsupportedFormat(
                "Legacy XLS format is not supported",
                stage="excel",
            ) from exc
        raise MalformedDocument(f"Cannot open Excel: {exc}", stage="excel") from exc

    try:
        for sheet_idx, sheet_name in enumerate(wb.sheetnames):
            check_timeout(start, lim.max_processing_seconds)
            if sheet_idx >= lim.max_sheets:
                break

            sheet = wb[sheet_name]
            rows = _extract_sheet_rows(sheet, sheet_name, lim)
            if rows:
                tables.append(
                    GenericTable(
                        rows=tuple(rows),
                        table_index=sheet_idx,
                        page_or_sheet=sheet_name,
                        provenance={"source_url": source_url} if source_url else {},
                    )
                )
    finally:
        wb.close()

    return ExtractionOutput(
        tables=tuple(tables),
        text_blocks=tuple(text_blocks),
        errors=tuple(errors),
        metadata={"sheet_count": len(tables)},
    )


def _extract_sheet_rows(
    sheet: Any,
    sheet_name: str,
    limits: ExtractionLimits,
) -> list[GenericRow]:
    rows: list[GenericRow] = []
    row_count = 0

    for row_idx, row in enumerate(sheet.iter_rows(max_row=limits.max_rows_per_table)):
        if row_count >= limits.max_rows_per_table:
            break
        row_count += 1

        cells: list[GenericCell] = []
        for col_idx, cell in enumerate(row):
            if col_idx >= limits.max_cols_per_table:
                break

            value = cell.value
            if value is None:
                text = ""
            elif isinstance(value, (int, float)):
                text = str(value)
            else:
                text = str(value).strip()

            truncated = limits.enforce_cell_length(text)
            cells.append(
                GenericCell(
                    text=truncated,
                    row_index=row_idx,
                    col_index=col_idx,
                    is_header=(row_idx == 0),
                )
            )

        if cells and any(c.text for c in cells):
            rows.append(
                GenericRow(
                    cells=tuple(cells),
                    row_index=row_idx,
                    is_header=(row_idx == 0),
                )
            )

    return rows
